#!/usr/bin/python

# manojmw
# 19 Apr, 2022

import argparse, sys
import openpyxl as xl
import scipy.stats as stats
import gzip
import logging
import re

###########################################################

# Parses tab-seperated canonical transcripts file
# Required columns are: 'ENSG' and 'GENE' (can be in any order,
# but they MUST exist)
#
# Returns a dictionary:
# - Key -> ENSG
# - Value -> Gene
def ENSG_Gene(inCanonicalFile):

    logging.info("Starting to run...")

    # Dictionary to store ENSG & Gene data
    ENSG_Gene_dict = {}

    # Opening canonical transcript file (gzip or non-gzip)
    try:
        if inCanonicalFile.endswith('.gz'):
            Canonical_File = gzip.open(inCanonicalFile, 'rt')
        else:
            Canonical_File = open(inCanonicalFile)
    except IOError:
        logging.error("At Step 5.2_addInteractome - Failed to read the Canonical transcript file: %s" % inCanonicalFile)
        sys.exit()

    Canonical_header_line = Canonical_File.readline() # Grabbing the header line

    Canonical_header_fields = Canonical_header_line.split('\t')

    # Check the column headers and grab indexes of our columns of interest
    (ENSG_index, Gene_index) = (-1,-1)

    for i in range(len(Canonical_header_fields)):
        if Canonical_header_fields[i] == 'ENSG':
            ENSG_index = i
        elif Canonical_header_fields[i] == 'GENE':
            Gene_index = i

    if not ENSG_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'ENSG' in the file: %s \n" % inCanonicalFile)
        sys.exit()
    elif not Gene_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'GENE' in the file: %s \n" % inCanonicalFile)
        sys.exit()
    # else grabbed the required column indexes -> PROCEED

    # Data lines
    for line in Canonical_File:
        line = line.rstrip('\n')
        CanonicalTranscripts_fields = line.split('\t')

        # Key -> ENSG
        # Value -> Gene

        ENSG_Gene_dict[CanonicalTranscripts_fields[ENSG_index]] = CanonicalTranscripts_fields[Gene_index]

    # Closing the file
    Canonical_File.close()

    return ENSG_Gene_dict

###########################################################

# Parses the candidateGenes file in .xlsx format
# Required columns are: 'Gene' & 'pathologyID' 
# (can be in any order, but they MUST exist)
# Also parses {ENSG_Gene_dict} returned
# by the function: ENSG_Gene
# Maps Candidate Gene names to ENSG
#
# Returns a dictionary
# - Key: ENSG of Candidate Gene
# - Value: list of pathology(s)
def CandidateGeneParser(inCandidateFile, ENSG_Gene_dict):
    
    # Input - List of candidate gene file(s)
    candidate_files = inCandidateFile

    # Dictionary to store candidate genes
    # and associated data
    # Key: Candidate Gene
    # Value: list of pathology(s)
    CandidateGene_dict = {}

    # Data lines
    for file in candidate_files:

        # Creating a workbook object 
        candF_wbobj = xl.load_workbook(file)

        # Creating a sheet object from the active attribute
        candF_sheetobj = candF_wbobj.active

        # Dictionary to store Candidate Gene and pathology data
        # Key -> rowindex of pathologyID
        # Value -> pathologyID
        Gene_patho_dict = {}
        
        # Iterating over col cells and checking if any header 
        # in the the header line matches our header of interest
        for header_cols in candF_sheetobj.iter_cols(1, candF_sheetobj.max_column):
            if header_cols[0].value == "pathologyID":
                for patho_field in header_cols[1:]:
                    # Skip empty fields
                    if patho_field.value == None:
                        pass
                    else:
                        Gene_patho_dict[patho_field.row] = patho_field.value

        # Grabbing gene names
        # Replacing the key (rowindex of pathologyID) in Gene_patho_dict
        # with a our new key (Gene_identifier)
        # Gene_identifier -> Gene name & rowindex of Gene name seperated by an '_'
        # This is to make sure that we do not replace the existsing gene and patho 
        # if the same gene is associated with a different pathology (in a given file,
        # row index will be unique)
        # Using a new for-loop because the keys in Gene_patho_dict will 
        # not be defined until we exit for loop
        # If key is not defined, then we cannot replace the old key with
        # our new Gene_identifier using the same row index
        for header_cols in candF_sheetobj.iter_cols(1, candF_sheetobj.max_column):
            if header_cols[0].value == "Gene":
                for Gene_field in header_cols[1:]:
                    # Skip empty fields
                    if Gene_field.value == None:
                        pass
                    else:
                        # Replacing the key in Gene_patho_dict with our new key (Gene_identifier)
                        Gene_patho_dict[Gene_field.value + '_' + str(Gene_field.row)] = Gene_patho_dict.pop(Gene_field.row)

        # List to store Gene name and pathology
        # We are not using the dictionary for further steps because
        # As we parse other candidate gene files, if the same gene (key)
        # is associated with a different pathology, the existing 
        # gene-pathology pair will be replaced as dictionary cannot 
        # contain redundant keys
        for Gene_identifier in Gene_patho_dict:
            Gene_identifierF = Gene_identifier.split('_')

            # Gene_identifierF[0] -> Gene name

            for ENSG in ENSG_Gene_dict.keys():
                if Gene_identifierF[0] == ENSG_Gene_dict[ENSG]:
                    Gene = ENSG
                    break
            
            Pathology = Gene_patho_dict[Gene_identifier]

            # Check if the Gene exists in CandidateGene_dict
            # Happens when same gene is associated with different pathology
            # If Gene exists, then append the new pathology to the list of pathologies
            if CandidateGene_dict.get(Gene, False):
                # Avoid adding same pathology more than once
                if not Pathology in CandidateGene_dict[Gene]:
                    CandidateGene_dict[Gene].append(Pathology)
            else:
                CandidateGene_dict[Gene] = [Pathology]
            
    return CandidateGene_dict

###########################################################

# Parses the patient samples metadata file
# Required column: 'pathologyID' 
# (can be in any order, but it MUST exist)
#
# Returns a list containing all the pathologies/Phenotypes
def getPathologies(inSample):

    sampleFile = inSample

    # List for storing pathologies
    pathologies_list = []

     # Creating a workbook object 
    sampF_wbobj = xl.load_workbook(sampleFile)

    # Creating a sheet object from the active attribute
    sampF_sheetobj = sampF_wbobj.active

    # Iterating over col cells and checking if any header 
    # in the the header line matches our header of interest
    for header_cols in sampF_sheetobj.iter_cols(1, sampF_sheetobj.max_column):
        if header_cols[0].value == "pathologyID":
            for patho_field in header_cols[1:]:
                # Skip empty fields
                if patho_field.value == None:
                    pass
                else:
                    if not patho_field.value in pathologies_list:
                        pathologies_list.append(patho_field.value)                                             

    return pathologies_list

###########################################################

# Parses the CandidateGene_dict & pathologies_list
#
# Counts the total number of candidate genes
# associated with each pathology
#
# Returns a list with the total count candidate genes (for each pathology)
def CountCandidateGenes(CandidateGene_dict, pathologies_list):

    # List for counting total candidate genes
    # associated with each pathology
    pathology_CandidateCount = [0] * len(pathologies_list)

    # Data lines
    for candidateGene in CandidateGene_dict:
        for pathology in CandidateGene_dict[candidateGene]:
            for i in range(len(pathologies_list)):
                if pathology == pathologies_list[i]:
                    pathology_CandidateCount[i] += 1              

    return pathology_CandidateCount

###########################################################

# Parses the High-quality Interactome produced by Interactome.py
#
# Required columns:
# First column - ENSG of Protein A
# Second column - ENSG of Protein B
#
# Returns 2 dictionaries and 1 list:
# First dictionary contains:
# - key: Protein A; Value: List of interactors
# Second dictionary contains:
# - key: Protein B; Value: List of interactors
# These dictionaries are later used to determine
# the no. of interactors for a given protein/gene
#
# The list contains all the interacting proteins from the interactome
def Interacting_Proteins(inInteractome):

    # Dictionaries to store interacting proteins
    # ProtA_dict - key: Protein A; Value: List of interactors
    # ProtB_dict - key: Protein B; Value: List of interactors
    ProtA_dict = {}
    ProtB_dict = {}

    # List of all interactors from the Interactome
    All_Interactors_list = []
        
    # Input - Interactome file
    Interactome_File = open(inInteractome)

    # Data lines
    for line in Interactome_File:
        line = line.rstrip('\n')

        Interactome_fields = line.split('\t')

        if Interactome_fields[0] != Interactome_fields[1]:
            # Check if the Key(ProtA) exists in ProtA_dict
            # If yes, then append the interctor to 
            # the list of values (Interactors)
            if ProtA_dict.get(Interactome_fields[0], False):
                ProtA_dict[Interactome_fields[0]].append(Interactome_fields[1])
            else:
                ProtA_dict[Interactome_fields[0]] = [Interactome_fields[1]]

            # Check if the Key(ProtB) exists in ProtB_dict
            # If yes, then append the interctor to 
            # the list of values (Interactors)
            if ProtB_dict.get(Interactome_fields[1], False):
                ProtB_dict[Interactome_fields[1]].append(Interactome_fields[0])
            else:
                ProtB_dict[Interactome_fields[1]] = [Interactome_fields[0]]    

            # Storing all the interactors in All_Interactors_list
            if not Interactome_fields[0] in All_Interactors_list:
                All_Interactors_list.append(Interactome_fields[0])
            elif not Interactome_fields[1] in All_Interactors_list:
                All_Interactors_list.append(Interactome_fields[1])
        # else:
            # NOOP -> The interaction is a self-interaction

    # Closing the file
    Interactome_File.close()

    return ProtA_dict, ProtB_dict, All_Interactors_list

###########################################################

 # Parses the UniProt Primary Accession file produced by Uniprot_parser.py
 # Required columns are: 'Primary_AC' and 'ENSGs' (can be in any order,
 # but they MUST exist)
 #
 # Also parses the dictionary ENSG_Gene_dict
 # returned by the function ENSG_Gene
 #
 # Maps UniProt Primary Accession to ENSG
 # Returns the count of UniProt accessions with unique ENSGs
 #
 # Count corresponds to total human genes
 # This count is later used for calculating
 # Benjamini-Hochberg adjusted P-values
def Uniprot_ENSG(inUniProt, ENSG_Gene_dict):

    # Counter for accessions with single canonical human ENSG
    Count_UniqueENSGs = 0

    Uniprot_File = open(inUniProt)

    # Grabbing the header line
    Uniprot_header = Uniprot_File.readline()

    Uniprot_header = Uniprot_header.rstrip('\n')

    Uniprot_header_fields = Uniprot_header.split('\t')

    # Check the column header and grab indexes of our columns of interest
    (UniProt_PrimAC_index, ENSG_index) = (-1, -1)

    for i in range(len(Uniprot_header_fields)):
        if Uniprot_header_fields[i] == 'Primary_AC':
            UniProt_PrimAC_index = i
        elif Uniprot_header_fields[i] == 'ENSGs':
            ENSG_index = i

    if not UniProt_PrimAC_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'Primary_AC' in the file: %s \n" % inUniProt)
        sys.exit()
    elif not ENSG_index >= 0:
        logging.error("At Step 5.2_addInteractome - Missing required column title 'ENSG' in the file: %s \n" % inUniProt)
        sys.exit()
    # else grabbed the required column indexes -> PROCEED

    # Data lines
    for line in Uniprot_File:
        line = line.rstrip('\n')
        Uniprot_fields = line.split('\t')

        # ENSG column  - This is a single string containing comma-seperated ENSGs
        # So we split it into a list that can be accessed later
        UniProt_ENSGs = Uniprot_fields[ENSG_index].split(',')
        
        canonical_human_ENSGs = []

        # If ENSG is in the canonical transcripts file
        # Append it to canonical_human_ENSGs
        for ENSG in UniProt_ENSGs:
            if ENSG in ENSG_Gene_dict.keys():
                canonical_human_ENSGs.append(ENSG)

        # Keeping the count of protein with single ENSGs
        if len(canonical_human_ENSGs) == 1:
            Count_UniqueENSGs += 1

    Uniprot_File.close()

    return Count_UniqueENSGs

###########################################################

# Cluster output file produced by clustering methods, 
# processed by ProcessClusterFile_*.py script (if required)
#
# Returns a dictionary:
# Key: clusterID as in the inFile
# Value: Dictionary containing 2 types of key/value pair:
# -> First type of key/value pair:
#   - Key: ENSG of the node;  
#   - Value: 1
# -> Second type of key/value pair:
#   - Key: Pathology name; 
#   - Value -> List containing P-value & candidate genes
#              present in the cluster for each pathology
#              
def Build_ClusterDict(inClusterFile, CandidateGene_dict, pathologies_list, pathology_CandidateCount, Count_UniqueENSGs):

    # Dictionary to store Clustering data 
    # and pathology-specific P-values
    IntCluster_dict = {}

    Cluster_File = open(inClusterFile)

    # Compiling regular expressions

    # Matching Cluster ID
    re_cluster = re.compile('(^ClusterID.\d+)\|\|$')

    # Matching ENSG of nodes
    re_node = re.compile('^(ENSG\d+)$')

    # Intializing variable to store
    # clusterIDS and nodes
    Clust_ID = ''
    ENSG_nodes = []

    # A list containing Distinct lists 
    # One sublist per pathology
    # Each sublist contains the count and names of candidate genes
    # present in the cluster
    # The first item of each sublist will be count which will be later
    # be replaced with cluster associated P-Value
    clusterCandidate = [[0] for i in range(len(pathologies_list))]

    # skip header
    Cluster_File.readline()

    # Data lines 
    for line in Cluster_File:

        line = line.rstrip("\r\n")

        # Get the cluster ID
        if (re_cluster.match(line)):
            Clust_ID = re_cluster.match(line).group(1)
        # Get the ENSG of the nodes in the current cluster
        elif (re_node.match(line)):
            ENSG = re_node.match(line).group(1)
            if not ENSG in ENSG_nodes:
                ENSG_nodes.append(ENSG)
        
        # End of the cluster is indicated by an empty line
        # If we reach here, process data for the current cluster
        elif (line == ''):
            # If the size of the cluster is < 3 
            # Then, do not store data for this cluster
            # Empty the accumulators and move on to next cluster
            if len(ENSG_nodes) < 3:
                # Reset the accumulators and move on to the next cluster
                Clust_ID = ''
                ENSG_nodes = []
                clusterCandidate = [[0] for i in range(len(pathologies_list))]
                continue
            elif len(ENSG_nodes) > 130:
                logging.error("The size of the cluster "+Clust_ID+" is > 130. Please fix the cluster file" )
                sys.exit()
                
            # If the size of the cluster is >= 2 and <=130
            else:
                # Storing ClusterID as the Key IntCluster_dict
                # Value is a dictionary containing 2 types of Key/value pair:
                # -> First type of key/value pair:
                #   - Key: ENSG of the node;  Value: 1
                # -> Second type of key/value pair:
                #   - Key: Pathology name;  P-value for the given cluster
            
                IntCluster_dict[Clust_ID] = {}

                # Storing the ENSG of each node as a key in IntCluster_dict[Clust_ID]
                # Value = 1
                for ENSG_node in ENSG_nodes:
                    if not ENSG_node in IntCluster_dict[Clust_ID]:
                        IntCluster_dict[Clust_ID][ENSG_node] = 1 

                # For the current cluster, checking how many
                # genes are known candidate genes and storing
                # the count in clusterCandidateCount
                for ENSG in IntCluster_dict[Clust_ID]:
                    if ENSG in CandidateGene_dict:
                        for patho in CandidateGene_dict[ENSG]:
                            for i in range(len(pathologies_list)):        
                                if patho == pathologies_list[i]:
                                    clusterCandidate[i][0] += 1
                                    clusterCandidate[i].append(ENSG)

                # Applying Fisher's exact test and computing P-values
                for i in range(len(clusterCandidate)):

                    # In the current cluster, if no gene is 
                    # a known candidate gene for the current pathology
                    # then there is no point in computing P-Values for this pathology
                    # So, we assign P-value = 1
                    if clusterCandidate[i] == 0:
                        cluster_p_value = 1

                    # If in the current cluster, gene(s) is a Known candidate
                    # gene(s) for the current Pathology, then we compute the P-value for 
                    # this pathology
                    else:
                        # Applying Fisher's exact test to calculate p-values
                        ComputePvalue_cluster = [[clusterCandidate[i][0], len(ENSG_nodes)],[pathology_CandidateCount[i], Count_UniqueENSGs]]
                        (odds_ratio, cluster_p_value) = stats.fisher_exact(ComputePvalue_cluster)
                        # Replacing the candidate gene count with the P-value
                        clusterCandidate[i][0] = cluster_p_value

                    # Storing Pathology name as the key in IntCluster_dict[Clust_ID]
                    # Value: List containing P-value & candidate genes
                    #        present in the cluster for the current pathology
                    IntCluster_dict[Clust_ID][pathologies_list[i]] = clusterCandidate[i]

                # Reset the accumulators and move on to the next cluster
                Clust_ID = ''
                ENSG_nodes = []
                clusterCandidate = [[0] for i in range(len(pathologies_list))]
                continue
    
    Cluster_File.close()

    return IntCluster_dict

###########################################################

# Parses a GTEX file
#
# Extracts the required data

# Returns a dictionary and a list
# Dictionary contains"
# - Key: ENSG (from the Gene ID column)
# - Value: List of GTEX favourite tissue ratios (calculated) 
#          and tissue-specific GTEX tpm values
# List contains newly built GTEX header
def getGTEX(inGTEXFile):

    GTEXFile = open(inGTEXFile)

    # A list to store favorite tissues
    # Working on Infertility here
    favouriteTissues = ['testis', 'ovary']

    favouriteTissIndex = [-1, -1]

    # Dicitionary to store GTEX data
    GTEX_dict = {}

    for line in GTEXFile:

        line = line.rstrip("\n\r")

        if line.startswith("#"): # Skip Comment lines
            continue

        elif line.startswith("Gene ID"): # Header line

            # next line is a header line
            GTEX_header_line = line

            # Sanity Check
            if not GTEX_header_line.startswith("Gene ID"):
                logging.error("Line should be GTEX header but can't parse it %s" % GTEX_header_line)
                sys.exit()
            else:
                tissues = GTEX_header_line.split('\t')
                tissues = tissues[2:]

            # Check the column headers and grab indexes of our columns of interest
            for i in range(len(tissues)):
                for fti in range(len(favouriteTissues)):
                    if tissues[i] == favouriteTissues[fti]:
                        if favouriteTissIndex[fti] != -1:
                            logging.error("Found favorite tissue %s twice" % tissues[i])
                            sys.exit()
                        else:
                            favouriteTissIndex[fti] = i
                            break 
                tissues[i] = tissues[i].replace(" ", "_")
                tissues[i] = 'GTEX_'+tissues[i]

            
            # FavTissues Index Sanity Check
            for fti in range(len(favouriteTissues)):
                if favouriteTissIndex[fti] == -1:
                    logging.error("Could not find favorite tissue %s column in header" % favouriteTissues[fti])
                    sys.exit()

        else:
            GTEX_data = line.split('\t')
            if not len(GTEX_data) == len(tissues) + 2:
                logging.error("Wrong number of fields in the line %s" % line)
                sys.exit()
            
            # Grab ENSG
            ENSG = GTEX_data.pop(0)

            if ENSG in GTEX_dict:
                logging.error("ENSG "+ENSG+" present twice in GTEX file %s" % line)
                sys.exit()

            # The second column is Gene name 
            # We want to ignore this column
            # So we remove it from GTEX_data
            GTEX_data.pop(0)

            # thisGTEX: list of strings holding expression values, one per tissue
            thisGTEX = [""] * len(tissues)

            # calculated GTEX ratio for each favorite tissue
            favTissRatios = [""] * len(favouriteTissues)

            # for calculating GTEX_*_RATIO:
            # sum of all GTEX values
            sumOfGtex = 0

            # number of defined GTEX values
            numberOfGtex = 0

            for i in range(len(GTEX_data)):
                if GTEX_data[i] == '':
                    thisGTEX[i] = 0
                    sumOfGtex += 0
                else:    
                    thisGTEX[i] = float(GTEX_data[i])
                    sumOfGtex += float(GTEX_data[i])
                numberOfGtex += 1

            # favExp / averageExp == favExp / (sumOfGtex / numberOfGtex) == favExp * numberOfGtex / sumOfGtex
            # so make sure we can divide by sumOfGtex
            if sumOfGtex == 0:
                logging.error("Sum of GTEX values is zero for gene "+ENSG+", impossible %s" % line)
                sys.exit()

            for ti in range(len(favouriteTissIndex)):
                favTissRatios[ti] = thisGTEX[favouriteTissIndex[ti]] * numberOfGtex / sumOfGtex

            # List to store GTEX_RATIOs first, then favorites, then others
            toPrint = []

            for favTissRatio in favTissRatios:
                # print max 2 digits after decimal
                favTR = round(favTissRatio, 2)
                toPrint.append(favTR)

            for fti in favouriteTissIndex:
                toPrint.append(thisGTEX[fti])
            
            for i in range(len(tissues)):
                if i in favouriteTissIndex:
                    continue
                else:
                    toPrint.append(thisGTEX[i])
            
            GTEX_dict[ENSG] = toPrint
    
    # List to store new GTEX header
    newGTEXHeader = []

    for ft in favouriteTissues:
        GTEXRatioheader = "GTEX_"+ft+"_RATIO"
        newGTEXHeader.append(GTEXRatioheader)

    for fti in favouriteTissIndex:
        newGTEXHeader.append(tissues[fti])
    
    for i in range(len(tissues)):
        if i in favouriteTissIndex:
            continue
        else:
            newGTEXHeader.append(tissues[i])
 
    GTEXFile.close()
        
    return GTEX_dict, newGTEXHeader
            
###########################################################

# Parses the dictionaries and list returned
# by the function: Interacting_Proteins
# Checks the number of interactors for each gene
# Checks the number of known interactors
# using the candidateGene_out_list returned by the function: CandidateGene2ENSG
#
# Prints to STDOUT in .tsv format
# The output consists of following data for each line (one gene per line):
#  -> Gene Name
#  -> If a gene is already a known candidate (adds the patho names - comma-separated)
#  -> Total Number of Interactors
#  -> Following information is added for each Pathology:
#       - Known Interactors
#       - List of Known Interactors
#       - Known Interactors P-value
#       - If a gene is 'PRESENT' in an Enriched Cluster
#       - ClusterID (if enriched)
#       - Size of the Cluster
#       - List of Candidate genes present in the Cluster
#       - The Cluster associated P-value
#       - Count of second degree neighbors that are Known candidates
#       - List of second degree neighbors that are Known candidates
# -> GTEX data
def Interactors_PValue(args):

    # Calling the functions
    ENSG_Gene_dict = ENSG_Gene(args.inCanonicalFile)
    CandidateGene_dict = CandidateGeneParser(args.inCandidateFile, ENSG_Gene_dict)
    pathologies_list = getPathologies(args.inSample)
    pathology_CandidateCount = CountCandidateGenes(CandidateGene_dict, pathologies_list)
    (ProtA_dict, ProtB_dict, All_Interactors_list) = Interacting_Proteins(args.inInteractome)
    Count_UniqueENSGs = Uniprot_ENSG(args.inUniProt, ENSG_Gene_dict)
    IntCluster_dict = Build_ClusterDict(args.inClusterFile, CandidateGene_dict, pathologies_list, pathology_CandidateCount, Count_UniqueENSGs)
    (GTEX_dict, newGTEXHeader) = getGTEX(args.inGTEXFile)

    # Printing header
    Patho_header_list = [[patho+'_INTERACTORS_COUNT', patho+'_INTERACTORS', patho+'_INTERACTORS_PVALUE', patho+'_ENRICHED_CLUSTER', patho+'_ENRICHED_CLUSTER_ID', patho+'_ENRICHED_CLUSTER_SIZE', patho+'_ENRICHED_CLUSTER_CANDIDATE_GENES', patho+'_ENRICHED_CLUSTER_PVALUE', patho+'_SECOND_DEGREE_INTERACTORS_COUNT', patho+'_SECOND_DEGREE_INTERACTORS'] for patho in pathologies_list]
    print('GENE' + '\t' + 'KNOWN_CANDIDATE_GENE' + '\t' + 'TOTAL_INTERACTORS' + '\t' + '\t'.join(header for Patho_headerIndex in range(len(Patho_header_list)) for header in Patho_header_list[Patho_headerIndex]) + '\t' + '\t'.join(GTEXHeader for GTEXHeader in newGTEXHeader))


    # Checking the number of interactors for each gene
    for ENSG_index in range(len(All_Interactors_list)):

        # Initializing a list to store data for the 
        # current gene 
        Gene_AllPatho_Pvalue = []

        Gene_AllPatho_Pvalue.append(All_Interactors_list[ENSG_index])
        
        Known_Pathology = []

        # Checking if the gene is a known candidate gene for any pathology
        if All_Interactors_list[ENSG_index] in CandidateGene_dict:
            for patho in CandidateGene_dict[All_Interactors_list[ENSG_index]]:
                Known_Pathology.append(patho)
        
        # Storing Known Known_Pathologies as a single comma seperated string
        Known_Pathologystr = ','.join(patho for patho in Known_Pathology)

        Gene_AllPatho_Pvalue.append(Known_Pathologystr)

        # List of interactors for the current gene
        Interactors = []

        # If Protein is the first protein
        if (All_Interactors_list[ENSG_index] in ProtA_dict.keys()):
            # Get the interacting protein
            for Interactor in ProtA_dict[All_Interactors_list[ENSG_index]]:
                if not Interactor in Interactors:
                    Interactors.append(Interactor)
                    
        # If Protein is the Second protein
        if (All_Interactors_list[ENSG_index] in ProtB_dict.keys()):
            # Get the interacting protein
            for Interactor in ProtB_dict[All_Interactors_list[ENSG_index]]:
                if not Interactor in Interactors:
                    Interactors.append(Interactor)

        Gene_AllPatho_Pvalue.append(len(Interactors))     

        for i in range(len(pathologies_list)):

            # List for known interactor(s)
            Known_Interactors = []

            # Initializing a list to store data for each pathology
            Output_eachPatho = []

            # List to store second degree neighbors 
            # that are known candidates
            AllsecondDegreeKnownInt = []

            # Checking if the interactor is a known ENSG (candidate ENSG)
            for interactor in Interactors:
                if interactor in CandidateGene_dict.keys():
                    for pathology in CandidateGene_dict[interactor]:
                        if pathology == pathologies_list[i]:
                            Known_Interactors.append(interactor)

                # List to store second degree interactors
                # seen in ProtA_dict
                secondDegreeInt_ProtA_dict = []

                # Checking if the second degree neighbours 
                # are known candidates in ProtA_dict
                if interactor in ProtA_dict.keys():
                    for secondDegreeInt in ProtA_dict[interactor]:
                        # Add every second degree interactor seen
                        # in ProtA_dict to secondDegreeInt_ProtA_dict
                        secondDegreeInt_ProtA_dict.append(secondDegreeInt)
                        if secondDegreeInt in CandidateGene_dict.keys():
                            for pathology in CandidateGene_dict[secondDegreeInt]:
                                if pathology == pathologies_list[i]:
                                    AllsecondDegreeKnownInt.append(ENSG_Gene_dict[secondDegreeInt])

                # If the interactor is also present in ProtB_dict
                if interactor in ProtB_dict.keys():
                    for secondDegreeInt in ProtB_dict[interactor]:
                        # But if the second degree interactor was not already
                        # seen in ProtA_dict for the current interactor
                        # then check if this second degree interactor is a candidate gene
                        if not secondDegreeInt in secondDegreeInt_ProtA_dict:
                            if secondDegreeInt in CandidateGene_dict.keys():
                                for pathology in CandidateGene_dict[secondDegreeInt]:
                                    if pathology == pathologies_list[i]:
                                        AllsecondDegreeKnownInt.append(ENSG_Gene_dict[secondDegreeInt])

            # Getting the Gene name for Known Interactors
            for Known_InteractorIndex in range(len(Known_Interactors)):
                Known_Interactors[Known_InteractorIndex] = ENSG_Gene_dict[Known_Interactors[Known_InteractorIndex]]

            if Known_Interactors:
                # Applying Fisher's exact test to calculate p-values
                ComputePvalue_data = [[len(Known_Interactors), len(Interactors)],[pathology_CandidateCount[i], Count_UniqueENSGs]]
                (odds_ratio, p_value) = stats.fisher_exact(ComputePvalue_data)

            # If there are no Known Interactors, 
            # there is no point is computing P-value,
            # So we assign P-value as 1
            else:
                p_value = 1

            if Known_Interactors:
                # Storing Known Interactors as a single comma seperated string
                Known_InteractorsStr = ','.join(Known_Interactor for Known_Interactor in Known_Interactors)
                Output_eachPatho = [len(Known_Interactors), Known_InteractorsStr, p_value]
            else:
                Output_eachPatho = [len(Known_Interactors), '', p_value]
            
            # Adding Interactome clustering data
            for cluster in IntCluster_dict:
                # Checking if the current gene is present in any cluster
                if All_Interactors_list[ENSG_index] in IntCluster_dict[cluster]:
                    #  If the gene is present in the cluster, get the P-value
                    #  for this cluster
                    clusterPatho_data = IntCluster_dict[cluster].get(pathologies_list[i])

                    # The first item of the clusterPatho_data will always be a P-value
                    clusterPatho_Pvalue = clusterPatho_data[0] 

                    # Get the names of candidate genes present in the cluster and store
                    # it as a single comma-seperated string
                    clusterPatho_KnownInteractorsstr = ','.join(ENSG_Gene_dict[Known_Interactor] for Known_Interactor in clusterPatho_data[1:])

                    # If P-value is not equal to 1 means, the cluster is enriched
                    # for the current pathology
                    # So we say this gene is PRESENT in the enriched cluster
                    # and add the P-value associated with this cluster for the current pathology
                    if clusterPatho_Pvalue != 1:
                        Output_eachPatho.append('PRESENT')
                        Output_eachPatho.append(cluster)
                        # We know that clusterID is the key and value is a dictionary
                        # containing 2 types of key/value pair (including pathology-specific information)
                        # So, size of the cluster will be excluding pathology information (i.e pathology key/value pair)
                        Cluster_size = len(IntCluster_dict[cluster]) - len(pathologies_list)
                        Output_eachPatho.append(Cluster_size)
                        Output_eachPatho.append(clusterPatho_KnownInteractorsstr)
                        Output_eachPatho.append(clusterPatho_Pvalue)
                        break

            # If the gene is not present in any cluster
            # This can happen as we eliminate clusters with size < 2
            # So we append empty values and assign P-value = 1
            # similar to a gene that is present in a cluster but not 
            # enriched for the current pathology
            if len(Output_eachPatho) == 3:
                no_cluster_data = ['', '', 0, '', 1]
                for empty_data in no_cluster_data:
                    Output_eachPatho.append(empty_data)
            else: # len(Output_eachPatho) is 5
                # The gene was present in one of the clusters that
                # is enriched for the current pathology
                pass

            # Appending all known interactors in the 
            # 2-hop neighborhood
            for Known_Interactor in Known_Interactors:
                AllsecondDegreeKnownInt.append(Known_Interactor)

            # Adding second degree known interactors data
            Output_eachPatho.append(len(AllsecondDegreeKnownInt))
            AllsecondDegreeKnownIntstr = ','.join(Known_Interactor for Known_Interactor in AllsecondDegreeKnownInt)
            Output_eachPatho.append(AllsecondDegreeKnownIntstr)

            for data in Output_eachPatho:
                Gene_AllPatho_Pvalue.append(data)

        # Adding GTEX Data
        if All_Interactors_list[ENSG_index] in GTEX_dict:
            for GTEX_value in GTEX_dict[All_Interactors_list[ENSG_index]]:
                Gene_AllPatho_Pvalue.append(GTEX_value)
        else: 
            pass 

        # Getting the Gene name for the ENSG
        Gene_AllPatho_Pvalue[0] = ENSG_Gene_dict[Gene_AllPatho_Pvalue[0]]

        print('\t'.join(str(data) for data in Gene_AllPatho_Pvalue))

    logging.info("All done, completed successfully!")

    return

###########################################################

# Taking and handling command-line arguments
def main():
    file_parser = argparse.ArgumentParser(description =
    """
--------------------------------------------------------------------------------------------------------------
Program: Parses the input files. For a each gene, adds the Interactome data (both Naive & Clustering approach) 
         assoicated with each pathology. Next adds the GTEX data and prints to STDOUT in .tsv format
--------------------------------------------------------------------------------------------------------------

The output contains following data for each gene (one gene per line):
 -> Gene Name
 -> If a gene is already a known candidate (adds the patho names - comma-separated)
 -> Total Number of Interactors
 -> Following information is added for each Pathology:
    - Known Interactors Count
    - List of Known Interactors (comma-separated)
    - Known Interactors P-value
    - If a gene is 'PRESENT' in an Enriched Cluster
    - ClusterID (if PRESENT)
    - Size of the Cluster
    - List of Candidate genes present in the Cluster (comma-separated)
    - The Cluster associated P-value
    - Count of second degree neighbors that are Known candidates
    - List of second degree neighbors that are Known candidates (comma-separated)
 -> GTEX data
--------------------------------------------------------------------------------------------------------------

Arguments [defaults] -> Can be abbreviated to shortest unambiguous prefixes
    """,
    formatter_class = argparse.RawDescriptionHelpFormatter)

    required = file_parser.add_argument_group('Required arguments')
    optional = file_parser.add_argument_group('Optional arguments')

    required.add_argument('--inSampleFile', metavar = "Input File", dest = "inSample", help = 'Sample metadata file', required=True)
    required.add_argument('--inUniprot', metavar = "Input File", dest = "inUniProt", help = 'Uniprot output File generated by the UniProt_parser.py', required=True)
    required.add_argument('--inCandidateFile', metavar = "Input File", dest = "inCandidateFile", nargs = '*', help = 'Candidate Genes Input File name(.xlsx)', required=True)
    required.add_argument('--inCanonicalFile', metavar = "Input File", dest = "inCanonicalFile", help = 'Canonical Transcripts file (.gz or non .gz)', required=True)
    required.add_argument('--inInteractome', metavar = "Input File", dest = "inInteractome", help = 'Input File Name (High-quality Human Interactome(.tsv) produced by Build_Interactome.py)', required=True)
    required.add_argument('--inClusterFile', metavar = "Input File", dest = "inClusterFile", help = 'Cluster output file produced by clustering methods (If required, processed by appropriate ProcessClusterFile_*.py script', required=True)
    required.add_argument('--inGTEXFile', metavar = "Input File", dest = "inGTEXFile", help = 'GTEX Input File name', required=True)

    args = file_parser.parse_args()
    Interactors_PValue(args)

if __name__ == "__main__":
    # Logging to Standard Error
    Log_Format = "%(levelname)s - %(asctime)s - %(message)s \n"
    logging.basicConfig(stream = sys.stderr, format  = Log_Format, level = logging.DEBUG)
    main()
