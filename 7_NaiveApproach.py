#!/usr/bin/python

# manojmw
# 17 Feb, 2022

import argparse, sys
import openpyxl as xl
import scipy.stats as stats
import gzip
import logging

###########################################################

# Parses tab-seperated canonical transcripts file
# Required columns are: 'ENSG' and 'GENE' (can be in any order,
# but they MUST exist)
#
# Returns a dictionary:
# - Key -> ENSG
# - Value -> Gene
def ENSG_Gene(inCanonicalFile):

    logging.info("Starting to run...")

    # Dictionary to store ENSG & Gene data
    ENSG_Gene_dict = {}

    # Excute below code only if Canonical Transcript file is provided
    if inCanonicalFile:

        logging.info("Processing data from Canonical Transcripts File: %s" % inCanonicalFile)

        # Opening canonical transcript file (gzip or non-gzip)
        try:
            if inCanonicalFile.endswith('.gz'):
                Canonical_File = gzip.open(inCanonicalFile, 'rt')
            else:
                Canonical_File = open(inCanonicalFile)
        except IOError:
            logging.error("At Step 5.2_addInteractome - Failed to read the Canonical transcript file: %s" % inCanonicalFile)
            sys.exit()

        Canonical_header_line = Canonical_File.readline() # Grabbing the header line

        Canonical_header_fields = Canonical_header_line.split('\t')

        # Check the column headers and grab indexes of our columns of interest
        (ENSG_index, Gene_index) = (-1,-1)

        for i in range(len(Canonical_header_fields)):
            if Canonical_header_fields[i] == 'ENSG':
                ENSG_index = i
            elif Canonical_header_fields[i] == 'GENE':
                Gene_index = i

        if not ENSG_index >= 0:
            logging.error("At Step 5.2_addInteractome - Missing required column title 'ENSG' in the file: %s \n" % inCanonicalFile)
            sys.exit()
        elif not Gene_index >= 0:
            logging.error("At Step 5.2_addInteractome - Missing required column title 'GENE' in the file: %s \n" % inCanonicalFile)
            sys.exit()
        # else grabbed the required column indexes -> PROCEED

        # Data lines
        for line in Canonical_File:
            line = line.rstrip('\n')
            CanonicalTranscripts_fields = line.split('\t')

            # Key -> ENSG
            # Value -> Gene

            ENSG_Gene_dict[CanonicalTranscripts_fields[ENSG_index]] = CanonicalTranscripts_fields[Gene_index]

        # Closing the file
        Canonical_File.close()

    return ENSG_Gene_dict

###########################################################

# Parses the candidateGenes file in .xlsx format
# Required columns are: 'Gene' & 'pathologyID' 
# (can be in any order, but they MUST exist)
# Also parses {ENSG_Gene_dict} returned
# by the function: ENSG_Gene
# Maps Candidate Gene names to ENSG
#
# Returns a dictionary
# - Key: ENSG of Candidate Gene
# - Value: list of pathology(s)
def CandidateGeneParser(inCandidateFile, ENSG_Gene_dict):
    
    # Input - List of candidate gene file(s)
    candidate_files = inCandidateFile

    # Dictionary to store candidate genes
    # and associated data
    # Key: Candidate Gene
    # Value: list of pathology(s)
    CandidateGene_dict = {}

    # Run the below code only if candidate 
    # gene file(s) is provided, else return
    # an empty dictionary
    if candidate_files:
        # Data lines
        for file in candidate_files:

            logging.info("Processing data from Candidate Gene File: %s" % file)

            # Creating a workbook object 
            candF_wbobj = xl.load_workbook(file)

            # Creating a sheet object from the active attribute
            candF_sheetobj = candF_wbobj.active

            # Dictionary to store Candidate Gene and pathology data
            # Key -> rowindex of pathologyID
            # Value -> pathologyID
            Gene_patho_dict = {}
            
            # Iterating over col cells and checking if any header 
            # in the the header line matches our header of interest
            for header_cols in candF_sheetobj.iter_cols(1, candF_sheetobj.max_column):
                if header_cols[0].value == "pathologyID":
                    for patho_field in header_cols[1:]:
                        # Skip empty fields
                        if patho_field.value == None:
                            pass
                        else:
                            Gene_patho_dict[patho_field.row] = patho_field.value

            # Grabbing gene names
            # Replacing the key (rowindex of pathologyID) in Gene_patho_dict
            # with a our new key (Gene_identifier)
            # Gene_identifier -> Gene name & rowindex of Gene name seperated by an '_'
            # This is to make sure that we do not replace the existsing gene and patho 
            # if the same gene is associated with a different pathology (in a given file,
            # row index will be unique)
            # Using a new for-loop because the keys in Gene_patho_dict will 
            # not be defined until we exit for loop
            # If key is not defined, then we cannot replace the old key with
            # our new Gene_identifier using the same row index
            for header_cols in candF_sheetobj.iter_cols(1, candF_sheetobj.max_column):
                if header_cols[0].value == "Gene":
                    for Gene_field in header_cols[1:]:
                        # Skip empty fields
                        if Gene_field.value == None:
                            pass
                        else:
                            # Replacing the key in Gene_patho_dict with our new key (Gene_identifier)
                            Gene_patho_dict[Gene_field.value + '_' + str(Gene_field.row)] = Gene_patho_dict.pop(Gene_field.row)

            # List to store Gene name and pathology
            # We are not using the dictionary for further steps because
            # As we parse other candidate gene files, if the same gene (key)
            # is associated with a different pathology, the existing 
            # gene-pathology pair will be replaced as dictionary cannot 
            # contain redundant keys
            for Gene_identifier in Gene_patho_dict:
                Gene_identifierF = Gene_identifier.split('_')

                # Gene_identifierF[0] -> Gene name

                for ENSG in ENSG_Gene_dict.keys():
                    if Gene_identifierF[0] == ENSG_Gene_dict[ENSG]:
                        Gene = ENSG
                        break
                
                Pathology = Gene_patho_dict[Gene_identifier]

                # Check if the Gene exists in CandidateGene_dict
                # Happens when same gene is associated with different pathology
                # If Gene exists, then append the new pathology to the list of pathologies
                if CandidateGene_dict.get(Gene, False):
                    # Avoid adding same pathology more than once
                    if not Pathology in CandidateGene_dict[Gene]:
                        CandidateGene_dict[Gene].append(Pathology)
                else:
                    CandidateGene_dict[Gene] = [Pathology]
            
    return CandidateGene_dict

###########################################################

# Parses the dictionary {CandidateGene_dict}
# returned by the function: CandidateGeneParser
#
# Returns a list containing all the pathologies/Phenotypes
def getPathologies(inSample):

    sampleFile = inSample

    logging.info("Processing data from Sample metadata File: %s" % sampleFile)

    # List for storing pathologies
    pathologies_list = []

     # Creating a workbook object 
    sampF_wbobj = xl.load_workbook(sampleFile)

    # Creating a sheet object from the active attribute
    sampF_sheetobj = sampF_wbobj.active

    # Iterating over col cells and checking if any header 
    # in the the header line matches our header of interest
    for header_cols in sampF_sheetobj.iter_cols(1, sampF_sheetobj.max_column):
        if header_cols[0].value == "pathologyID":
            for patho_field in header_cols[1:]:
                # Skip empty fields
                if patho_field.value == None:
                    pass
                else:
                    if not patho_field.value in pathologies_list:
                        pathologies_list.append(patho_field.value)                                             

    return pathologies_list

###########################################################

# Parses the CandidateGene_dict & pathologies_list
#
# Counts the total number of candidate genes
# associated with each pathology
#
# Returns a list with the total count candidate genes (for each pathology)
def CountCandidateGenes(CandidateGene_dict, pathologies_list):

    # List for counting total candidate genes
    # associated with each pathology
    pathology_CandidateCount = [0] * len(pathologies_list)

    # Data lines
    for candidateGene in CandidateGene_dict:
        for pathology in CandidateGene_dict[candidateGene]:
            for i in range(len(pathologies_list)):
                if pathology == pathologies_list[i]:
                    pathology_CandidateCount[i] += 1              

    return pathology_CandidateCount

###########################################################

# Parses the High-quality Interactome produced by Interactome.py
#
# Required columns:
# First column - ENSG of Protein A
# Second column - ENSG of Protein B
#
# Returns 2 dictionaries and 1 list:
# First dictionary contains:
# - key: Protein A; Value: List of interactors
# Second dictionary contains:
# key: Protein B; Value: List of interactors
# These dictionaries are later used to determine
# the no. of interactors for a given protein/gene
#
# The list contains all the interacting proteins from the interactome
def Interacting_Proteins(inInteractome):

    # Dictionaries to store interacting proteins
    # In ProtA_dict, key -> Protein A; Value -> Protein B
    # In ProtB_dict, key -> Protein B; Value -> Protein A
    ProtA_dict = {}
    ProtB_dict = {}

    # List of all interactors from the Interactome
    All_Interactors_list = []

    # Excute below code only if Interactome file is provided
    if inInteractome:
        
        # Input - Interactome file
        Interactome_File = open(inInteractome)

        logging.info("Processing data from Interactome File: %s" % inInteractome)

        # Data lines
        for line in Interactome_File:
            line = line.rstrip('\n')

            Interactome_fields = line.split('\t')

            if Interactome_fields[0] != Interactome_fields[1]:
                # Check if the Key(ProtA) exists in ProtA_dict
                # If yes, then append the interctor to 
                # the list of values (Interactors)
                if ProtA_dict.get(Interactome_fields[0], False):
                    ProtA_dict[Interactome_fields[0]].append(Interactome_fields[1])
                else:
                    ProtA_dict[Interactome_fields[0]] = [Interactome_fields[1]]

                # Check if the Key(ProtB) exists in ProtB_dict
                # If yes, then append the interctor to 
                # the list of values (Interactors)
                if ProtB_dict.get(Interactome_fields[1], False):
                    ProtB_dict[Interactome_fields[1]].append(Interactome_fields[0])
                else:
                    ProtB_dict[Interactome_fields[1]] = [Interactome_fields[0]]    

                # Storing all the interactors in All_Interactors_list
                if not Interactome_fields[0] in All_Interactors_list:
                    All_Interactors_list.append(Interactome_fields[0])
                elif not Interactome_fields[1] in All_Interactors_list:
                    All_Interactors_list.append(Interactome_fields[1])
            # else:
                # NOOP -> The interaction is a self-interaction

        # Closing the file
        Interactome_File.close()

    return ProtA_dict, ProtB_dict, All_Interactors_list

###########################################################

 # Parses the UniProt Primary Accession file produced by Uniprot_parser.py
 # Required columns are: 'Primary_AC' and 'ENSGs' (can be in any order,
 # but they MUST exist)
 #
 # Also parses the dictionary ENSG_Gene_dict
 # returned by the function ENSG_Gene
 #
 # Maps UniProt Primary Accession to ENSG
 # Returns the count of UniProt accessions with unique ENSGs
 #
 # Count corresponds to total human genes
 # This count is later used for calculating
 # Benjamini-Hochberg adjusted P-values
def Uniprot_ENSG(inUniProt, ENSG_Gene_dict):

    # Counter for accessions with single canonical human ENSG
    Count_UniqueENSGs = 0

    # Excute below code only if UniProt file is provided
    if inUniProt:

        Uniprot_File = open(inUniProt)

        # Grabbing the header line
        Uniprot_header = Uniprot_File.readline()

        Uniprot_header = Uniprot_header.rstrip('\n')

        Uniprot_header_fields = Uniprot_header.split('\t')

        # Check the column header and grab indexes of our columns of interest
        (UniProt_PrimAC_index, ENSG_index) = (-1, -1)

        for i in range(len(Uniprot_header_fields)):
            if Uniprot_header_fields[i] == 'Primary_AC':
                UniProt_PrimAC_index = i
            elif Uniprot_header_fields[i] == 'ENSGs':
                ENSG_index = i

        if not UniProt_PrimAC_index >= 0:
            logging.error("At Step 5.2_addInteractome - Missing required column title 'Primary_AC' in the file: %s \n" % inUniProt)
            sys.exit()
        elif not ENSG_index >= 0:
            logging.error("At Step 5.2_addInteractome - Missing required column title 'ENSG' in the file: %s \n" % inUniProt)
            sys.exit()
        # else grabbed the required column indexes -> PROCEED

        # Data lines
        for line in Uniprot_File:
            line = line.rstrip('\n')
            Uniprot_fields = line.split('\t')

            # ENSG column  - This is a single string containing comma-seperated ENSGs
            # So we split it into a list that can be accessed later
            UniProt_ENSGs = Uniprot_fields[ENSG_index].split(',')
            
            canonical_human_ENSGs = []

            # If ENSG is in the canonical transcripts file
            # Append it to canonical_human_ENSGs
            for ENSG in UniProt_ENSGs:
                if ENSG in ENSG_Gene_dict.keys():
                    canonical_human_ENSGs.append(ENSG)

            # Keeping the count of protein with single ENSGs
            if len(canonical_human_ENSGs) == 1:
                Count_UniqueENSGs += 1

    Uniprot_File.close()

    return Count_UniqueENSGs


###########################################################

# Parses the dictionaries and list returned
# by the function: Interacting_Proteins
# Checks the number of interactors for each gene
# Checks the number of known interactors
# using the candidateGene_out_list returned by the function: CandidateGene2ENSG
#
# Prints to STDOUT in .tsv format
# The output consists of following data foreach line:
# - Gene Name
# - Total Number of Interactors
# - Known Interactors, list of Known_Interactors & P-value for each Pathology
def Interactors_PValue(args):

    # Calling the functions
    # Calling the functions
    ENSG_Gene_dict = ENSG_Gene(args.inCanonicalFile)
    CandidateGene_dict = CandidateGeneParser(args.inCandidateFile, ENSG_Gene_dict)
    pathologies_list = getPathologies(args.inSample)
    pathology_CandidateCount = CountCandidateGenes(CandidateGene_dict, pathologies_list)
    (ProtA_dict, ProtB_dict, All_Interactors_list) = Interacting_Proteins(args.inInteractome)
    Count_UniqueENSGs = Uniprot_ENSG(args.inUniProt, ENSG_Gene_dict)

    # Initializing first output list containing distinct lists
    # i.e. one Sublist per Gene
    # Each Sublist contains:
    # - Gene name
    # - Total number of Interactors
    # - Known Interactors count, list of Known Interactors, P-value (for each pathology)
    Gene_AllPatho_Pvalue = [[] for i in range(len(All_Interactors_list))]

    logging.info("Processing data, Checking Interactors and Computing P-values...")

    # Checking the number of interactors for each gene
    for ENSG_index in range(len(All_Interactors_list)):

        Gene_AllPatho_Pvalue[ENSG_index].append(All_Interactors_list[ENSG_index])

        # List of interactors
        Interactors = []

        # If Protein is the first protein
        if (All_Interactors_list[ENSG_index] in ProtA_dict.keys()):
            # Get the interacting protein
            for Interactor in ProtA_dict[All_Interactors_list[ENSG_index]]:
                if not Interactor in Interactors:
                    Interactors.append(Interactor)
                    
        # If Protein is the Second protein
        if (All_Interactors_list[ENSG_index] in ProtB_dict.keys()):
            # Get the interacting protein
            for Interactor in ProtB_dict[All_Interactors_list[ENSG_index]]:
                if not Interactor in Interactors:
                    Interactors.append(Interactor)

        Gene_AllPatho_Pvalue[ENSG_index].append(len(Interactors))     

        for i in range(len(pathologies_list)):

            # List for known interactor(s)
            Known_Interactors = []

            # Initializing a list to store data for each pathology
            Output_eachPatho = []

            # Checking if the interactor is a known ENSG (candidate ENSG)
            for interactor in Interactors:
                if interactor in CandidateGene_dict.keys():
                    for pathology in CandidateGene_dict[interactor]:
                        if pathology == pathologies_list[i]:
                            Known_Interactors.append(interactor)

            # Getting the Gene name for Known Interactors
            for Known_InteractorIndex in range(len(Known_Interactors)):
                Known_Interactors[Known_InteractorIndex] = ENSG_Gene_dict[Known_Interactors[Known_InteractorIndex]]

            if Known_Interactors:
                # Applying Fisher's exact test to calculate p-values
                ComputePvalue_data = [[len(Known_Interactors), len(Interactors)],[pathology_CandidateCount[i], Count_UniqueENSGs]]
                (odds_ratio, p_value) = stats.fisher_exact(ComputePvalue_data)

            # If there are no Known Interactors, 
            # there is no point is computing P-value,
            # So we assign P-value as 1
            else:
                p_value = 1

            if Known_Interactors:
                # Storing Known Interactors as a single comma seperated string
                Known_InteractorsStr = ','.join(Known_Interactor for Known_Interactor in Known_Interactors)
                Output_eachPatho = [len(Known_Interactors), Known_InteractorsStr, p_value]
            else:
                Output_eachPatho = [len(Known_Interactors), '', p_value]

            for data in Output_eachPatho:
                Gene_AllPatho_Pvalue[ENSG_index].append(data)

        # Getting the Gene name for the ENSG
        Gene_AllPatho_Pvalue[ENSG_index][0] = ENSG_Gene_dict[Gene_AllPatho_Pvalue[ENSG_index][0]]

    # Printing header
    Patho_header_list = [[patho+'_KnownInteractorsCount', patho+'_KnownInteractors', patho+'_Pvalue'] for patho in pathologies_list]
    print('Gene\t', 'Total_Interactors\t', '\t'.join(header for Patho_headerIndex in range(len(Patho_header_list)) for header in Patho_header_list[Patho_headerIndex]))

    for Gene_AllPathoIndex in range(len(Gene_AllPatho_Pvalue)):
        print('\t'.join(str(eachGene_AllPatho_data) for eachGene_AllPatho_data in Gene_AllPatho_Pvalue[Gene_AllPathoIndex]))

    logging.info("All done, completed successfully!")

    return

###########################################################

# Taking and handling command-line arguments
def main():
    file_parser = argparse.ArgumentParser(description =
    """
-----------------------------------------------------------------------------------------------------------------------
Program: Parses the Interactome file generated by Interactome.py, checks the number of interactors for each gene. Next,
         parses the patient Candidate Gene file(s) and Canonical Transcripts file. Checks the number of interactors
         that are known candidate genes, computes P-values and prints to STDOUT in .tsv format
-----------------------------------------------------------------------------------------------------------------------
The output consists of following data for each line (one gene per line) :
 -> Gene Name
 -> Total Number of Interactors
 -> Known Interactors, list of Known Interactors & P-value for each Pathology
-----------------------------------------------------------------------------------------------------------------------

Arguments [defaults] -> Can be abbreviated to shortest unambiguous prefixes
    """,
    formatter_class = argparse.RawDescriptionHelpFormatter)

    required = file_parser.add_argument_group('Required arguments')
    optional = file_parser.add_argument_group('Optional arguments')

    required.add_argument('--inSampleFile', metavar = "Input File", dest = "inSample", help = 'Sample metadata file', required=True)
    required.add_argument('--inUniProt', metavar = "Input File", dest = "inUniProt", help = 'Uniprot output File generated by the UniProt_parser.py')
    required.add_argument('--inCandidateFile', metavar = "Input File", dest = "inCandidateFile", nargs = '*', help = 'Candidate Genes Input File name(.xlsx)')
    required.add_argument('--inCanonicalFile', metavar = "Input File", dest = "inCanonicalFile", help = 'Canonical Transcripts file (.gz or non .gz)')
    required.add_argument('--inInteractome', metavar = "Input File", dest = "inInteractome", help = 'Input File Name (High-quality Human Interactome(.tsv) produced by Build_Interactome.py)')

    args = file_parser.parse_args()
    Interactors_PValue(args)

if __name__ == "__main__":
    # Logging to Standard Error
    Log_Format = "%(levelname)s - %(asctime)s - %(message)s \n"
    logging.basicConfig(stream = sys.stderr, format  = Log_Format, level = logging.DEBUG)
    main()
